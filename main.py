from selenium import webdriver
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options
from bs4 import BeautifulSoup
import json
import time
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border

x = input("Platform 1 (pc, psn, xbox, switch): ")
y = input("Platform 2 (pc, psn, xbox, switch): ")
service = EdgeService(executable_path="edgedriver_win64/msedgedriver.exe")

options = Options()
#options.add_argument("headless")
options.add_argument("--start-fullscreen");


url1 = ("https://rl.insider.gg/en/"+x)
url2 = ("https://rl.insider.gg/en/"+y)
driver = webdriver.Edge(options = options, service=service)
print(f"Downloading {x} prices")
driver.get(url1)
time.sleep(10)
html = driver.page_source
soup1 = BeautifulSoup(html, "html.parser")

print(f"Downloading {y} prices")
driver.get(url2)
time.sleep(10)
html2 = driver.page_source
soup2 = BeautifulSoup(html2, "html.parser")
driver.quit()


#platform_2 =

#soup2 = BeautifulSoup(platform_1.content, "html.parser")

class Item:
    def __init__(self):
        self.name = None
        self.paint = None
        self.price = 0


def get_paint(i):
    x = i % 15
    match x:
        case 1:
            return "default"
        case 2:
            return "black"
        case 3:
            return "white"
        case 4:
            return "grey"
        case 5:
            return "crimson"
        case 6:
            return "pink"
        case 7:
            return "cobalt"
        case 8:
            return "skyblue"
        case 9:
            return "burnt"
        case 10:
            return "saffron"
        case 11:
            return "lime"
        case 12:
            return "green"
        case 13:
            return "orange"
        case 14:
            return "purple"
workbook = Workbook()
sheet = workbook.active
bold = Font(bold=True)

sheet["A1"] = "Item name"
sheet["A1"].font = bold
sheet["B1"] = "Paint"
sheet["B1"].font = bold
sheet["C1"] = f"Price {x}"
sheet["C1"].font = bold
sheet["D1"] = f"Price {y}"
sheet["D1"].font = bold
sheet["E1"] = "Dif"
sheet["E1"].font = bold
sheet["F1"] = "%"
sheet["F1"].font = bold



def write(item, i, color):

    if color == PatternFill(start_color='000000', end_color='000000', fill_type='solid'):
        font = Font(color='ffffff')
    else:
        font = Font(color='000000')

    sheet[f"A{i}"] = item.name
    sheet[f"A{i}"].font = font
    sheet[f"A{i}"].fill = color
    sheet[f"B{i}"] = item.paint
    sheet[f"B{i}"].font = font
    sheet[f"B{i}"].fill = color
    sheet[f"C{i}"] = item.price
    sheet[f"C{i}"].font = font
    sheet[f"C{i}"].fill = color

def write2(item, i, color):
    if color == PatternFill(start_color='000000', end_color='000000', fill_type='solid'):
        font = Font(color='ffffff')
    else:
        font = Font(color='000000')
    sheet[f"D{i}"] = item.price
    sheet[f"E{i}"] = float(sheet[f"C{i}"].value) - item.price
    try:
        sheet[f"F{i}"] = float(sheet[f"C{i}"].value) / item.price
    except ZeroDivisionError:
        sheet[f"F{i}"] = 0
    sheet[f"F{i}"].fill = color
    sheet[f"F{i}"].font = font
    sheet[f"D{i}"].fill = color
    sheet[f"D{i}"].font = font
    sheet[f"E{i}"].fill = color
    sheet[f"E{i}"].font = font
def real_price(price):
    #print(price)
    price = price.split()
    if 'k' in price:
        multiplier = 1000
    elif 'm' in price:
        multiplier = 1000000
    else:
        multiplier = 1
    try:
        a = float(price[0])*multiplier
        print(a)
        return a
    except (ValueError, IndexError) as e:
        print(price)
        print(e)
        return 0


def get_color(paint):
    match paint:
        case "default":
            return PatternFill(start_color='cac3b8', end_color='cac3b8', fill_type='solid')
        case "black":
            return PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        case "white":
            return PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')
        case "grey":
            return PatternFill(start_color='999999', end_color='999999', fill_type='solid')
        case "crimson":
            return PatternFill(start_color='ff6363', end_color='ff6363', fill_type='solid')
        case "pink":
            return PatternFill(start_color='f89eff', end_color='f89eff', fill_type='solid')
        case "cobalt":
            return PatternFill(start_color='506ec9', end_color='506ec9', fill_type='solid')
        case "skyblue":
            return PatternFill(start_color='63ffff', end_color='63ffff', fill_type='solid')
        case "burnt":
            return PatternFill(start_color='b46f45', end_color='b46f45', fill_type='solid')
        case "saffron":
            return PatternFill(start_color='ffff63', end_color='ffff63', fill_type='solid')
        case "lime":
            return PatternFill(start_color='63ff63', end_color='63ff63', fill_type='solid')
        case "green":
            return PatternFill(start_color='457337', end_color='457337', fill_type='solid')
        case "orange":
            return PatternFill(start_color='ffaa63', end_color='ffaa63', fill_type='solid')
        case "purple":
            return PatternFill(start_color='a862fc', end_color='a862fc', fill_type='solid')


def painted_bms(soup1, soup2):
    results = soup1.find(id="paintedBMDecalsPrices")
    price_elements = results.find("tbody").find_all("tr")
    x = 0
    for items in price_elements:
        i = 14
        #print(item)
        for paint in items.find_all("td"):
            x += 1
            i += 1
            price = paint.text
            if price != '-' and i % 15 == 0:
                name = price[:int(len(price)/2)]
                item = Item()
                item.name = name
            else:
                item.price = real_price(price)
                item.paint = get_paint(i)
                color = get_color(item.paint)
                write(item, x+1, color)

    results2 = soup2.find(id="paintedBMDecalsPrices")
    price_elements2 = results2.find("tbody").find_all("tr")
    x = 0
    for items in price_elements2:
        i = 14
        #print(item)
        for paint in items.find_all("td"):
            x += 1
            i += 1
            price = paint.text
            if price != '-' and i % 15 == 0:
                name = price[:int(len(price)/2)]
                item = Item()
                item.name = name
            else:
                item.price = real_price(price)
                item.paint = get_paint(i)
                color = get_color(item.paint)
                write2(item, x+1, color)
        return i


def painted_ges(soup1, soup2, x):
    y = x
    results = soup1.find(id="paintedGoalExplosionsPrices")
    price_elements = results.find("tbody").find_all("tr")
    for items in price_elements:
        i = 14
        #print(item)
        for paint in items.find_all("td"):
            x += 1
            i += 1
            price = paint.text
            if price != '-' and i % 15 == 0:
                name = price[:int(len(price)/2)]
                item = Item()
                item.name = name
            else:
                item.price = real_price(price)
                item.paint = get_paint(i)
                color = get_color(item.paint)
                write(item, x+1, color)

    results2 = soup2.find(id="paintedGoalExplosionsPrices")
    price_elements2 = results2.find("tbody").find_all("tr")
    x = y
    for items in price_elements2:
        i = 14
        #print(item)
        for paint in items.find_all("td"):
            x += 1
            i += 1
            price = paint.text
            if price != '-' and i % 15 == 0:
                name = price[:int(len(price)/2)]
                item = Item()
                item.name = name
            else:
                item.price = real_price(price)
                item.paint = get_paint(i)
                color = get_color(item.paint)
                print(item, x+1, color)
        return x
i = painted_bms(soup1, soup2)
#i = painted_ges(soup1, soup2, i)
workbook.save(filename="prices.xlsx")